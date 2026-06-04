# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl"]
# ///
"""
Create a professionally formatted Excel workbook with investment banking
standard styling.

Usage:
    uv run scripts/create_formatted_excel.py [output_path]

This is a reusable template. Adapt the data section for your use case.
"""

import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter


# ── Color Palette (Investment Banking Standard) ──────────────────────

# Fonts
BLUE_FONT = Font(color="0000FF", size=10, name="Calibri")
BLUE_FONT_BOLD = Font(color="0000FF", size=10, name="Calibri", bold=True)
BLACK_FONT = Font(color="000000", size=10, name="Calibri")
BLACK_FONT_BOLD = Font(color="000000", size=10, name="Calibri", bold=True)
GREEN_FONT = Font(color="008000", size=10, name="Calibri")
GREEN_FONT_BOLD = Font(color="008000", size=10, name="Calibri", bold=True)
WHITE_FONT_BOLD = Font(color="FFFFFF", size=10, name="Calibri", bold=True)
HEADER_FONT = Font(color="FFFFFF", size=12, name="Calibri", bold=True)
TITLE_FONT = Font(color="1F4E79", size=14, name="Calibri", bold=True)
SUBTITLE_FONT = Font(color="404040", size=10, name="Calibri", italic=True)

# Fills
DARK_BLUE_FILL = PatternFill("solid", fgColor="4472C4")
LIGHT_BLUE_FILL = PatternFill("solid", fgColor="D9E1F2")
INPUT_GREEN_FILL = PatternFill("solid", fgColor="E2EFDA")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
LIGHT_GRAY_FILL = PatternFill("solid", fgColor="F2F2F2")

# Sensitivity gradient fills (manual, for when conditional formatting isn't suitable)
SENS_DEEP_RED = PatternFill("solid", fgColor="F4CCCC")
SENS_LIGHT_RED = PatternFill("solid", fgColor="FCE4D6")
SENS_NEUTRAL = PatternFill("solid", fgColor="FFF2CC")
SENS_LIGHT_GREEN = PatternFill("solid", fgColor="D9EAD3")
SENS_DEEP_GREEN = PatternFill("solid", fgColor="B6D7A8")

# Borders
THIN_BORDER = Border(bottom=Side(style="thin", color="B2B2B2"))
BOTTOM_MEDIUM = Border(bottom=Side(style="medium", color="000000"))
BOTTOM_DOUBLE = Border(bottom=Side(style="double", color="000000"))
ALL_THIN = Border(
    left=Side(style="thin", color="B2B2B2"),
    right=Side(style="thin", color="B2B2B2"),
    top=Side(style="thin", color="B2B2B2"),
    bottom=Side(style="thin", color="B2B2B2"),
)

# Alignment
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


# ── Helper Functions ─────────────────────────────────────────────────

def apply_header_row(ws, row, labels, start_col=1):
    """Apply dark blue header styling to a row of labels."""
    for i, label in enumerate(labels):
        cell = ws.cell(row=row, column=start_col + i, value=label)
        cell.font = WHITE_FONT_BOLD
        cell.fill = DARK_BLUE_FILL
        cell.alignment = CENTER


def apply_data_row(ws, row, values, start_col=1, font=None, number_format=None,
                   fill=None, border=None):
    """Write a row of values with consistent formatting."""
    font = font or BLACK_FONT
    for i, val in enumerate(values):
        cell = ws.cell(row=row, column=start_col + i, value=val)
        cell.font = font
        if number_format:
            cell.number_format = number_format
        if fill:
            cell.fill = fill
        if border:
            cell.border = border
        cell.alignment = RIGHT if isinstance(val, (int, float)) else LEFT


def apply_input_cell(ws, row, col, value, number_format=None):
    """Style a cell as user input (blue font, green fill)."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = BLUE_FONT
    cell.fill = INPUT_GREEN_FILL
    if number_format:
        cell.number_format = number_format
    return cell


def add_sensitivity_table(ws, start_row, start_col, row_header, col_header,
                          row_values, col_values, data_matrix):
    """
    Create a sensitivity table with conditional formatting.

    Args:
        ws: Worksheet
        start_row/start_col: Top-left corner of the table
        row_header/col_header: Labels for the axes
        row_values: List of values for rows (e.g., WACC rates)
        col_values: List of values for columns (e.g., growth rates)
        data_matrix: 2D list of result values
    """
    # Column header label
    ws.cell(row=start_row, column=start_col + 1, value=col_header).font = BLACK_FONT_BOLD

    # Column values
    for j, cv in enumerate(col_values):
        cell = ws.cell(row=start_row, column=start_col + 1 + j, value=cv)
        cell.font = BLUE_FONT_BOLD
        cell.alignment = CENTER

    # Row header label
    ws.cell(row=start_row + 1, column=start_col, value=row_header).font = BLACK_FONT_BOLD

    # Data cells
    for i, rv in enumerate(row_values):
        # Row label
        cell = ws.cell(row=start_row + 1 + i, column=start_col, value=rv)
        cell.font = BLUE_FONT_BOLD
        cell.alignment = CENTER

        for j, dv in enumerate(data_matrix[i]):
            cell = ws.cell(row=start_row + 1 + i, column=start_col + 1 + j, value=dv)
            cell.font = BLACK_FONT
            cell.number_format = '$#,##0'
            cell.alignment = CENTER
            cell.border = ALL_THIN

    # Apply conditional formatting (red-yellow-green gradient)
    data_range = (
        f"{get_column_letter(start_col + 1)}{start_row + 1}:"
        f"{get_column_letter(start_col + len(col_values))}{start_row + len(row_values)}"
    )
    rule = ColorScaleRule(
        start_type="min", start_color="F8696B",
        mid_type="percentile", mid_value=50, mid_color="FFEB84",
        end_type="max", end_color="63BE7B",
    )
    ws.conditional_formatting.add(data_range, rule)


def auto_column_widths(ws, min_width=10, max_width=20):
    """Auto-adjust column widths based on content.

    CJK characters occupy ~2 character widths in Excel, so we count them
    as 2 instead of 1 to avoid truncated columns.
    """
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                s = str(cell.value)
                # CJK chars (U+4E00–U+9FFF, fullwidth, etc.) occupy ~2 widths
                width = sum(2 if '\u4e00' <= c <= '\u9fff' or
                            '\u3000' <= c <= '\u303f' or
                            '\uff00' <= c <= '\uffef' else 1
                            for c in s)
                max_len = max(max_len, width)
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))


# ── Example: Create a DCF Summary ───────────────────────────────────

def create_example_workbook(output_path: str):
    """Create an example professionally formatted Excel workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = "DCF Summary"

    # Title
    ws.cell(row=1, column=1, value="DCF Valuation Summary").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Example Company — Base Case").font = SUBTITLE_FONT

    # Key assumptions header
    apply_header_row(ws, 4, ["Parameter", "Value", "Source"])

    # Key assumptions data
    assumptions = [
        ("WACC", 0.10, "Calculated"),
        ("Terminal Growth Rate", 0.03, "Assumption"),
        ("Shares Outstanding (M)", 2580, "10-K Filing"),
        ("Net Debt ($M)", 28000, "Balance Sheet"),
    ]
    for i, (param, value, source) in enumerate(assumptions):
        r = 5 + i
        ws.cell(row=r, column=1, value=param).font = BLACK_FONT
        apply_input_cell(ws, r, 2, value,
                         number_format='0.0%' if isinstance(value, float) and value < 1 else '#,##0')
        ws.cell(row=r, column=3, value=source).font = GREEN_FONT

    # Separator
    for col in range(1, 4):
        ws.cell(row=9, column=col).border = BOTTOM_MEDIUM

    # Valuation output
    ws.cell(row=10, column=1, value="Implied Share Price").font = BLACK_FONT_BOLD
    cell = ws.cell(row=10, column=2, value=580)
    cell.font = BLACK_FONT_BOLD
    cell.number_format = '$#,##0'
    cell.border = BOTTOM_DOUBLE

    # Sensitivity table
    ws.cell(row=12, column=1, value="Sensitivity Analysis").font = TITLE_FONT

    wacc_values = [0.08, 0.09, 0.10, 0.11, 0.12]
    growth_values = [0.01, 0.02, 0.03, 0.04, 0.05]
    # Example data matrix (WACC rows x Growth cols)
    data_matrix = [
        [720, 780, 850, 940, 1050],
        [640, 690, 740, 800, 870],
        [570, 610, 650, 700, 750],
        [510, 540, 580, 620, 660],
        [460, 490, 520, 550, 580],
    ]

    add_sensitivity_table(
        ws, start_row=14, start_col=1,
        row_header="WACC", col_header="Terminal Growth Rate",
        row_values=wacc_values, col_values=growth_values,
        data_matrix=data_matrix,
    )

    # Format WACC/growth as percentages
    for r in range(15, 20):
        ws.cell(row=r, column=1).number_format = '0.0%'
    for c in range(2, 7):
        ws.cell(row=14, column=c).number_format = '0.0%'

    auto_column_widths(ws)
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Created: {output_path}")


if __name__ == "__main__":
    output = sys.argv[1] if len(sys.argv) > 1 else "example_output.xlsx"
    create_example_workbook(output)
